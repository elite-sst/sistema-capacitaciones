import os
from io import BytesIO

import pandas as pd
import streamlit as st

from sqlalchemy import text
from utils.db import engine

from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo

from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from openpyxl.utils import get_column_letter


# =========================================================
# RENDER ADMIN
# =========================================================

def render_admin():

    # =========================================================
    # HEADER ERP
    # =========================================================

    st.html("""

        <div class="banner">

            <div class="banner-left">

                <div class="banner-title">
                    🔐 Panel Administrador
                </div>

                <div class="banner-sub">
                    Gestión corporativa Elite Ingenieros
                </div>

            </div>

            <div class="banner-badge">
                Sistema Seguro
            </div>

        </div>

        """)

    # =========================================================
    # LOGIN ADMIN
    # =========================================================

    ADMIN_PASSWORD = st.secrets["ADMIN_PASSWORD"]

    if "admin_autenticado" not in st.session_state:

        st.session_state.admin_autenticado = False

    if not st.session_state.admin_autenticado:

        col1, col2, col3 = st.columns([1, 1.3, 1])

        with col2:

            st.markdown(
                """
                <div class="login-glow"></div>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                """
                <div class="login-top-icon">
                    🛡️
                </div>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                """
                <div class="login-title">
                    Acceso Corporativo
                </div>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                """
                <div class="login-subtitle">
                    Plataforma segura de administración empresarial
                </div>
                """,
                unsafe_allow_html=True
            )

            password = st.text_input(
                "Contraseña",
                type="password",
                label_visibility="collapsed",
                placeholder="Ingrese contraseña administrador",
                key="login_password"
            )

            st.markdown(
                "<div style='height:18px'></div>",
                unsafe_allow_html=True
            )

            if st.button(
                "Ingresar al Sistema",
                use_container_width=True,
                key="login_button"
            ):

                if password == ADMIN_PASSWORD:

                    st.session_state.admin_autenticado = True

                    st.rerun()

                else:

                    st.error(
                        "❌ Contraseña incorrecta"
                    )

            st.markdown(
                """
                <div class="login-footer">
                    Elite Ingenieros · ERP SaaS
                </div>
                """,
                unsafe_allow_html=True
            )

            st.stop()

    # =========================================================
    # TABS
    # =========================================================

    tab_formacion, tab_empleados, tab_asistencias = st.tabs(
        [
            "📚 Crear Formación",
            "👥 Gestión Empleados",
            "📋 Asistencias y Reportes"
        ]
    )

        # =========================================================
    # TAB FORMACIÓN
    # =========================================================

    with tab_formacion:

        st.subheader("📚 Crear Formación")

        subtab_crear_formacion, subtab_consultar_formacion, subtab_consolidado_formacion = st.tabs(
            [
                "➕ Crear",
                "🔎 Consultar / Editar",
                "📆 Consolidado Mensual"
            ]
        )

        # =====================================================
        # SUBTAB CREAR
        # =====================================================

        with subtab_crear_formacion:

            col1, col2 = st.columns(2)

            with col1:

                nombre_formacion = st.text_input(
                    "Nombre capacitación"
                )

            with col2:

                fecha = st.date_input(
                    "Fecha asistencia"
                )

            formador = st.text_input(
                "Formador",
                value="Eduardo Florez"
            )

            # =====================================================
            # TIPO REGISTRO
            # =====================================================

            tipo_registro = st.radio(
                "Tipo de registro",
                [
                    "Charla",
                    "Capacitación"
                ],
                horizontal=True
            )

            # =====================================================
            # VARIABLES PREGUNTAS
            # =====================================================

            pregunta_1 = None
            pregunta_2 = None
            pregunta_3 = None
            pregunta_4 = None
            pregunta_5 = None

            p1_opcion_a = None
            p1_opcion_b = None
            p1_opcion_c = None
            p1_opcion_d = None
            p1_correcta = None

            p2_opcion_a = None
            p2_opcion_b = None
            p2_opcion_c = None
            p2_opcion_d = None
            p2_correcta = None

            p3_opcion_a = None
            p3_opcion_b = None
            p3_opcion_c = None
            p3_opcion_d = None
            p3_correcta = None

            p4_opcion_a = None
            p4_opcion_b = None
            p4_opcion_c = None
            p4_opcion_d = None
            p4_correcta = None

            p5_opcion_a = None
            p5_opcion_b = None
            p5_opcion_c = None
            p5_opcion_d = None
            p5_correcta = None

            # =====================================================
            # PREGUNTAS CAPACITACIÓN
            # =====================================================

            if tipo_registro == "Capacitación":

                st.markdown("### 📝 Preguntas de evaluación")

                for i in range(1, 6):

                    with st.expander(
                        f"Pregunta {i}",
                        expanded=(i == 1)
                    ):

                        globals()[f"pregunta_{i}"] = st.text_input(
                            f"Texto pregunta {i}",
                            key=f"crear_pregunta_{i}"
                        )

                        col_a, col_b = st.columns(2)

                        with col_a:

                            globals()[f"p{i}_opcion_a"] = st.text_input(
                                f"Pregunta {i} - Opción A",
                                key=f"crear_p{i}a"
                            )

                            globals()[f"p{i}_opcion_b"] = st.text_input(
                                f"Pregunta {i} - Opción B",
                                key=f"crear_p{i}b"
                            )

                        with col_b:

                            globals()[f"p{i}_opcion_c"] = st.text_input(
                                f"Pregunta {i} - Opción C",
                                key=f"crear_p{i}c"
                            )

                            globals()[f"p{i}_opcion_d"] = st.text_input(
                                f"Pregunta {i} - Opción D",
                                key=f"crear_p{i}d"
                            )

                        globals()[f"p{i}_correcta"] = st.radio(
                            f"Respuesta correcta pregunta {i}",
                            ["A", "B", "C", "D"],
                            horizontal=True,
                            key=f"crear_correcta_{i}"
                        )

            # =====================================================
            # CREAR FORMACIÓN
            # =====================================================

            if st.button(
                "Crear formación",
                use_container_width=True
            ):

                if not nombre_formacion:

                    st.warning(
                        "⚠️ Debe ingresar el nombre"
                    )

                elif (
                    tipo_registro == "Capacitación"
                    and
                    not globals()["pregunta_1"]
                ):

                    st.warning(
                        "⚠️ Para una capacitación debe ingresar al menos la pregunta 1."
                    )

                else:

                    try:

                        with engine.begin() as conn:

                            resultado = conn.execute(
                                text("""
                                    INSERT INTO formaciones (
                                        nombre_formacion,
                                        fecha_asistencia,
                                        formador,
                                        tipo_registro,

                                        pregunta_1,
                                        pregunta_2,
                                        pregunta_3,
                                        pregunta_4,
                                        pregunta_5,

                                        p1_opcion_a,
                                        p1_opcion_b,
                                        p1_opcion_c,
                                        p1_opcion_d,
                                        p1_correcta,

                                        p2_opcion_a,
                                        p2_opcion_b,
                                        p2_opcion_c,
                                        p2_opcion_d,
                                        p2_correcta,

                                        p3_opcion_a,
                                        p3_opcion_b,
                                        p3_opcion_c,
                                        p3_opcion_d,
                                        p3_correcta,

                                        p4_opcion_a,
                                        p4_opcion_b,
                                        p4_opcion_c,
                                        p4_opcion_d,
                                        p4_correcta,

                                        p5_opcion_a,
                                        p5_opcion_b,
                                        p5_opcion_c,
                                        p5_opcion_d,
                                        p5_correcta
                                    )
                                    VALUES (
                                        :nombre_formacion,
                                        :fecha_asistencia,
                                        :formador,
                                        :tipo_registro,

                                        :pregunta_1,
                                        :pregunta_2,
                                        :pregunta_3,
                                        :pregunta_4,
                                        :pregunta_5,

                                        :p1_opcion_a,
                                        :p1_opcion_b,
                                        :p1_opcion_c,
                                        :p1_opcion_d,
                                        :p1_correcta,

                                        :p2_opcion_a,
                                        :p2_opcion_b,
                                        :p2_opcion_c,
                                        :p2_opcion_d,
                                        :p2_correcta,

                                        :p3_opcion_a,
                                        :p3_opcion_b,
                                        :p3_opcion_c,
                                        :p3_opcion_d,
                                        :p3_correcta,

                                        :p4_opcion_a,
                                        :p4_opcion_b,
                                        :p4_opcion_c,
                                        :p4_opcion_d,
                                        :p4_correcta,

                                        :p5_opcion_a,
                                        :p5_opcion_b,
                                        :p5_opcion_c,
                                        :p5_opcion_d,
                                        :p5_correcta
                                    )
                                    RETURNING id
                                """),
                                {
                                    "nombre_formacion": nombre_formacion.strip(),
                                    "fecha_asistencia": fecha,
                                    "formador": formador.strip(),
                                    "tipo_registro": tipo_registro,

                                    **{
                                        f"pregunta_{i}": globals()[f"pregunta_{i}"]
                                        for i in range(1, 6)
                                    },

                                    **{
                                        f"p{i}_opcion_a": globals()[f"p{i}_opcion_a"]
                                        for i in range(1, 6)
                                    },

                                    **{
                                        f"p{i}_opcion_b": globals()[f"p{i}_opcion_b"]
                                        for i in range(1, 6)
                                    },

                                    **{
                                        f"p{i}_opcion_c": globals()[f"p{i}_opcion_c"]
                                        for i in range(1, 6)
                                    },

                                    **{
                                        f"p{i}_opcion_d": globals()[f"p{i}_opcion_d"]
                                        for i in range(1, 6)
                                    },

                                    **{
                                        f"p{i}_correcta": globals()[f"p{i}_correcta"]
                                        for i in range(1, 6)
                                    }
                                }
                            )

                            id_formacion = resultado.fetchone()[0]

                        url = (
                            f"https://elite-sst.streamlit.app/"
                            f"?view=asistencia&formacion={id_formacion}"
                        )

                        st.success(
                            "✅ Formación creada correctamente"
                        )

                        st.code(url)

                    except Exception as e:

                        st.error(
                            f"❌ Error: {e}"
                        )



        # =====================================================
        # SUBTAB CONSULTAR / EDITAR
        # =====================================================

        with subtab_consultar_formacion:

            st.subheader("🔎 Consultar / Editar Formación")

            with engine.begin() as conn:

                df_historial = pd.read_sql(
                    text("""
                        SELECT
                            id,
                            fecha_asistencia,
                            nombre_formacion,
                            tipo_registro,
                            formador
                        FROM formaciones
                        ORDER BY id DESC
                    """),
                    conn
                )

            if df_historial.empty:

                st.warning("⚠️ No existen formaciones registradas.")

            else:

                df_historial["url"] = (
                    "https://elite-sst.streamlit.app/"
                    "?view=asistencia&formacion="
                    + df_historial["id"].astype(str)
                )

                st.dataframe(
                    df_historial,
                    use_container_width=True
                )

                opciones_formacion = {
                    f"{row.id} - {row.nombre_formacion}": row.id
                    for row in df_historial.itertuples()
                }

                seleccion_formacion = st.selectbox(
                    "Seleccione formación para editar",
                    list(opciones_formacion.keys()),
                    key="editar_formacion_select"
                )

                id_formacion_editar = opciones_formacion[
                    seleccion_formacion
                ]

                url_formacion = (
                    "https://elite-sst.streamlit.app/"
                    f"?view=asistencia&formacion={id_formacion_editar}"
                )

                st.code(url_formacion)

                with engine.begin() as conn:

                    formacion = conn.execute(
                        text("""
                            SELECT *
                            FROM formaciones
                            WHERE id = :id
                        """),
                        {
                            "id": id_formacion_editar
                        }
                    ).mappings().first()

                if formacion:

                    st.markdown("---")

                    nombre_edit = st.text_input(
                        "Nombre formación",
                        value=formacion["nombre_formacion"] or "",
                        key=f"edit_nombre_{id_formacion_editar}"
                    )

                    fecha_edit = st.date_input(
                        "Fecha asistencia",
                        value=formacion["fecha_asistencia"],
                        key=f"edit_fecha_{id_formacion_editar}"
                    )

                    formador_edit = st.text_input(
                        "Formador",
                        value=formacion["formador"] or "",
                        key=f"edit_formador_{id_formacion_editar}"
                    )

                    tipo_edit = st.radio(
                        "Tipo de registro",
                        [
                            "Charla",
                            "Capacitación"
                        ],
                        horizontal=True,
                        index=0 if formacion["tipo_registro"] == "Charla" else 1,
                        key=f"edit_tipo_{id_formacion_editar}"
                    )

                    preguntas_edit = {}

                    if tipo_edit == "Capacitación":

                        st.markdown("### 📝 Preguntas de evaluación")

                        for i in range(1, 6):

                            correcta_actual = (
                                formacion[f"p{i}_correcta"]
                                if formacion[f"p{i}_correcta"] in ["A", "B", "C", "D"]
                                else "A"
                            )

                            with st.expander(
                                f"Pregunta {i}",
                                expanded=(i == 1)
                            ):

                                pregunta = st.text_input(
                                    f"Texto pregunta {i}",
                                    value=formacion[f"pregunta_{i}"] or "",
                                    key=f"edit_pregunta_{id_formacion_editar}_{i}"
                                )

                                col_a, col_b = st.columns(2)

                                with col_a:

                                    opcion_a = st.text_input(
                                        f"Pregunta {i} - Opción A",
                                        value=formacion[f"p{i}_opcion_a"] or "",
                                        key=f"edit_p{id_formacion_editar}_{i}_a"
                                    )

                                    opcion_b = st.text_input(
                                        f"Pregunta {i} - Opción B",
                                        value=formacion[f"p{i}_opcion_b"] or "",
                                        key=f"edit_p{id_formacion_editar}_{i}_b"
                                    )

                                with col_b:

                                    opcion_c = st.text_input(
                                        f"Pregunta {i} - Opción C",
                                        value=formacion[f"p{i}_opcion_c"] or "",
                                        key=f"edit_p{id_formacion_editar}_{i}_c"
                                    )

                                    opcion_d = st.text_input(
                                        f"Pregunta {i} - Opción D",
                                        value=formacion[f"p{i}_opcion_d"] or "",
                                        key=f"edit_p{id_formacion_editar}_{i}_d"
                                    )

                                correcta = st.radio(
                                    f"Respuesta correcta pregunta {i}",
                                    ["A", "B", "C", "D"],
                                    horizontal=True,
                                    index=["A", "B", "C", "D"].index(correcta_actual),
                                    key=f"edit_correcta_{id_formacion_editar}_{i}"
                                )

                                preguntas_edit[i] = {
                                    "pregunta": pregunta,
                                    "a": opcion_a,
                                    "b": opcion_b,
                                    "c": opcion_c,
                                    "d": opcion_d,
                                    "correcta": correcta
                                }

                    if st.button(
                        "💾 Guardar Cambios",
                        use_container_width=True,
                        key=f"guardar_cambios_{id_formacion_editar}"
                    ):

                        try:

                            valores = {
                                "id": id_formacion_editar,
                                "nombre_formacion": nombre_edit.strip(),
                                "fecha_asistencia": fecha_edit,
                                "formador": formador_edit.strip(),
                                "tipo_registro": tipo_edit
                            }

                            for i in range(1, 6):

                                if tipo_edit == "Capacitación":

                                    valores[f"pregunta_{i}"] = preguntas_edit[i]["pregunta"]
                                    valores[f"p{i}_a"] = preguntas_edit[i]["a"]
                                    valores[f"p{i}_b"] = preguntas_edit[i]["b"]
                                    valores[f"p{i}_c"] = preguntas_edit[i]["c"]
                                    valores[f"p{i}_d"] = preguntas_edit[i]["d"]
                                    valores[f"p{i}_correcta"] = preguntas_edit[i]["correcta"]

                                else:

                                    valores[f"pregunta_{i}"] = None
                                    valores[f"p{i}_a"] = None
                                    valores[f"p{i}_b"] = None
                                    valores[f"p{i}_c"] = None
                                    valores[f"p{i}_d"] = None
                                    valores[f"p{i}_correcta"] = None

                            with engine.begin() as conn:

                                conn.execute(
                                    text("""
                                        UPDATE formaciones
                                        SET
                                            nombre_formacion = :nombre_formacion,
                                            fecha_asistencia = :fecha_asistencia,
                                            formador = :formador,
                                            tipo_registro = :tipo_registro,

                                            pregunta_1 = :pregunta_1,
                                            pregunta_2 = :pregunta_2,
                                            pregunta_3 = :pregunta_3,
                                            pregunta_4 = :pregunta_4,
                                            pregunta_5 = :pregunta_5,

                                            p1_opcion_a = :p1_a,
                                            p1_opcion_b = :p1_b,
                                            p1_opcion_c = :p1_c,
                                            p1_opcion_d = :p1_d,
                                            p1_correcta = :p1_correcta,

                                            p2_opcion_a = :p2_a,
                                            p2_opcion_b = :p2_b,
                                            p2_opcion_c = :p2_c,
                                            p2_opcion_d = :p2_d,
                                            p2_correcta = :p2_correcta,

                                            p3_opcion_a = :p3_a,
                                            p3_opcion_b = :p3_b,
                                            p3_opcion_c = :p3_c,
                                            p3_opcion_d = :p3_d,
                                            p3_correcta = :p3_correcta,

                                            p4_opcion_a = :p4_a,
                                            p4_opcion_b = :p4_b,
                                            p4_opcion_c = :p4_c,
                                            p4_opcion_d = :p4_d,
                                            p4_correcta = :p4_correcta,

                                            p5_opcion_a = :p5_a,
                                            p5_opcion_b = :p5_b,
                                            p5_opcion_c = :p5_c,
                                            p5_opcion_d = :p5_d,
                                            p5_correcta = :p5_correcta

                                        WHERE id = :id
                                    """),
                                    valores
                                )

                            st.success("✅ Formación actualizada correctamente.")
                            st.info("🔗 La URL sigue siendo la misma porque depende del ID.")

                        except Exception as e:

                            st.error(
                                f"❌ Error actualizando formación: {e}"
                            )
        with subtab_consolidado_formacion:

            st.subheader("📆 Consolidado Mensual")

            col1, col2 = st.columns(2)

            with col1:
                fecha_inicio = st.date_input(
                    "Fecha inicial",
                    key="fecha_inicio_consolidado"
                )

            with col2:
                fecha_fin = st.date_input(
                    "Fecha final",
                    key="fecha_fin_consolidado"
                )

            with engine.begin() as conn:

                df_consolidado = pd.read_sql(
                    text("""
                        SELECT
                            f.id AS id_formacion,
                            f.fecha_asistencia,
                            f.nombre_formacion,
                            f.tipo_registro,
                            f.formador,
                            a.cedula,
                            a.nombre_completo,
                            a.cargo,
                            a.proyecto,
                            a.zona,
                            a.clasificacion_formacion,
                            a.tipo_formacion,
                            a.autoriza_datos,
                            a.respuesta_1,
                            a.resultado_1,
                            a.respuesta_2,
                            a.resultado_2,
                            a.respuesta_3,
                            a.resultado_3,
                            a.respuesta_4,
                            a.resultado_4,
                            a.respuesta_5,
                            a.resultado_5,
                            a.puntaje,
                            a.fecha_registro
                        FROM formaciones f
                        LEFT JOIN asistencias a
                            ON a.id_formacion = f.id
                        WHERE f.fecha_asistencia BETWEEN :fecha_inicio AND :fecha_fin
                        ORDER BY
                            f.fecha_asistencia DESC,
                            f.id DESC,
                            a.fecha_registro DESC
                    """),
                    conn,
                    params={
                        "fecha_inicio": fecha_inicio,
                        "fecha_fin": fecha_fin
                    }
                )

            if df_consolidado.empty:

                st.warning(
                    "⚠️ No hay formaciones registradas en ese rango."
                )

            else:

                df_resumen = (
                    df_consolidado
                    .groupby(
                        [
                            "id_formacion",
                            "fecha_asistencia",
                            "nombre_formacion",
                            "tipo_registro",
                            "formador"
                        ],
                        dropna=False
                    )
                    .agg(
                        total_asistencias=("cedula", "count")
                    )
                    .reset_index()
                )

                st.dataframe(
                    df_resumen,
                    use_container_width=True
                )

                opciones_consolidado = {
                    f"{row.id_formacion} - {row.nombre_formacion}": row.id_formacion
                    for row in df_resumen.itertuples()
                }

                seleccion_consolidado = st.selectbox(
                    "Seleccione formación para ver detalle",
                    list(opciones_consolidado.keys()),
                    key="select_consolidado"
                )

                id_consolidado = opciones_consolidado[seleccion_consolidado]

                fila_detalle = df_resumen[
                    df_resumen["id_formacion"] == id_consolidado
                ].iloc[0]

                st.markdown("---")
                st.markdown("### 📌 Detalle de la formación")

                col1, col2, col3, col4 = st.columns(4)

                with col1:
                    st.metric("ID", fila_detalle["id_formacion"])

                with col2:
                    st.metric("Tipo", fila_detalle["tipo_registro"])

                with col3:
                    st.metric("Fecha", str(fila_detalle["fecha_asistencia"]))

                with col4:
                    st.metric("Asistencias", int(fila_detalle["total_asistencias"]))

                st.write(
                    f"**Nombre:** {fila_detalle['nombre_formacion']}"
                )

                st.write(
                    f"**Formador:** {fila_detalle['formador']}"
                )

                # =====================================================
                # EXPORTAR CONSOLIDADO EXCEL PRO
                # =====================================================

                output = BytesIO()

                with pd.ExcelWriter(
                    output,
                    engine="openpyxl"
                ) as writer:

                    df_resumen.to_excel(
                        writer,
                        index=False,
                        sheet_name="Resumen_Formaciones"
                    )

                    df_consolidado.to_excel(
                        writer,
                        index=False,
                        sheet_name="Detalle_Asistencias"
                    )

                    for sheet_name in [
                        "Resumen_Formaciones",
                        "Detalle_Asistencias"
                    ]:

                        worksheet = writer.sheets[sheet_name]

                        # Encabezado verde corporativo
                        for cell in worksheet[1]:

                            cell.font = Font(
                                bold=True,
                                color="FFFFFF"
                            )

                            cell.alignment = Alignment(
                                horizontal="center",
                                vertical="center"
                            )

                            cell.fill = PatternFill(
                                start_color="166534",
                                end_color="166534",
                                fill_type="solid"
                            )

                        # Ajustar ancho de columnas
                        for column_cells in worksheet.columns:

                            length = max(
                                len(str(cell.value))
                                if cell.value else 0
                                for cell in column_cells
                            )

                            worksheet.column_dimensions[
                                get_column_letter(
                                    column_cells[0].column
                                )
                            ].width = length + 5

                        # Crear tabla estructurada Excel
                        total_filas = worksheet.max_row
                        total_columnas = worksheet.max_column

                        rango = (
                            f"A1:"
                            f"{get_column_letter(total_columnas)}"
                            f"{total_filas}"
                        )

                        nombre_tabla = (
                            "TablaResumenFormaciones"
                            if sheet_name == "Resumen_Formaciones"
                            else "TablaDetalleAsistencias"
                        )

                        tabla = Table(
                            displayName=nombre_tabla,
                            ref=rango
                        )

                        estilo = TableStyleInfo(
                            name="TableStyleMedium9",
                            showRowStripes=True,
                            showColumnStripes=False
                        )

                        tabla.tableStyleInfo = estilo

                        worksheet.add_table(tabla)

                output.seek(0)

                st.download_button(
                    label="📥 Descargar consolidado mensual Excel",
                    data=output,
                    file_name=(
                        f"Consolidado_Formaciones_"
                        f"{fecha_inicio}_{fecha_fin}.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet"
                    ),
                    use_container_width=True
                )
    
    # =========================================================
    # TAB EMPLEADOS
    # =========================================================

    with tab_empleados:

        st.subheader("👥 Gestión Empleados")

        subtab_agregar, subtab_consultar, subtab_estado = st.tabs(
            [
                "➕ Agregar",
                "🔎 Consultar",
                "⚙️ Activar / Inactivar"
            ]
        )

        # =====================================================
        # AGREGAR
        # =====================================================

        with subtab_agregar:

            col1, col2 = st.columns(2)

            with col1:
                cedula_emp = st.text_input("Cédula")
                nombre_emp = st.text_input("Nombre")

            with col2:
                cargo_emp = st.text_input("Cargo")
                zona_emp = st.text_input(
                    "Zona",
                    value="Metropolitano"
                )

            proyecto_emp = st.text_input(
                "Proyecto",
                value="CONEXIÓN Y VINCULACIÓN METROPOLITANO"
            )

            if st.button(
                "Guardar empleado",
                use_container_width=True
            ):

                if not cedula_emp or not nombre_emp or not cargo_emp:

                    st.warning("⚠️ Campos obligatorios")

                else:

                    try:

                        with engine.begin() as conn:

                            conn.execute(
                                text("""
                                    INSERT INTO empleados (
                                        cedula,
                                        nombre_completo,
                                        cargo,
                                        proyecto,
                                        zona,
                                        estado
                                    )
                                    VALUES (
                                        :cedula,
                                        :nombre_completo,
                                        :cargo,
                                        :proyecto,
                                        :zona,
                                        'ACTIVO'
                                    )
                                """),
                                {
                                    "cedula": cedula_emp.strip(),
                                    "nombre_completo": nombre_emp.strip(),
                                    "cargo": cargo_emp.strip(),
                                    "proyecto": proyecto_emp.strip(),
                                    "zona": zona_emp.strip()
                                }
                            )

                        st.success("✅ Empleado guardado")

                    except Exception as e:

                        st.error(f"❌ Error: {e}")

        # =====================================================
        # CONSULTAR
        # =====================================================

        with subtab_consultar:

            buscar = st.text_input(
                "Buscar empleado",
                key="buscar_empleado_consulta"
            )

            with engine.begin() as conn:

                df_empleados = pd.read_sql(
                    text("""
                        SELECT
                            cedula,
                            nombre_completo,
                            cargo,
                            proyecto,
                            zona,
                            estado,
                            fecha_creacion
                        FROM empleados
                        WHERE
                            :buscar = ''
                            OR cedula ILIKE :patron
                            OR nombre_completo ILIKE :patron
                            OR cargo ILIKE :patron
                        ORDER BY nombre_completo
                    """),
                    conn,
                    params={
                        "buscar": buscar.strip(),
                        "patron": f"%{buscar.strip()}%"
                    }
                )

            st.dataframe(
                df_empleados,
                use_container_width=True
            )

        # =====================================================
        # ACTIVAR / INACTIVAR
        # =====================================================

        with subtab_estado:

            st.markdown("### ⚙️ Activar / Inactivar Empleado")
           
            if "mensaje_estado_empleado" in st.session_state:

                tipo_mensaje, texto_mensaje = st.session_state["mensaje_estado_empleado"]

                if tipo_mensaje == "success":

                    st.success(texto_mensaje)

                elif tipo_mensaje == "warning":

                    st.warning(texto_mensaje)

                del st.session_state["mensaje_estado_empleado"]

            buscar_estado = st.text_input(
                "Buscar por cédula o nombre",
                key="buscar_empleado_estado"
            )

            with engine.begin() as conn:

                df_estado = pd.read_sql(
                    text("""
                        SELECT
                            cedula,
                            nombre_completo,
                            cargo,
                            proyecto,
                            zona,
                            estado
                        FROM empleados
                        WHERE
                            :buscar = ''
                            OR cedula ILIKE :patron
                            OR nombre_completo ILIKE :patron
                        ORDER BY nombre_completo
                    """),
                    conn,
                    params={
                        "buscar": buscar_estado.strip(),
                        "patron": f"%{buscar_estado.strip()}%"
                    }
                )

            if df_estado.empty:

                st.warning("⚠️ No se encontraron empleados.")

            else:

                st.dataframe(
                    df_estado,
                    use_container_width=True
                )

                opciones_empleados = {
                    f"{row.cedula} - {row.nombre_completo}": row.cedula
                    for row in df_estado.itertuples()
                }

                seleccion_empleado = st.selectbox(
                    "Seleccione empleado",
                    list(opciones_empleados.keys()),
                    key="select_estado_empleado"
                )

                cedula_seleccionada = opciones_empleados[
                    seleccion_empleado
                ]

                empleado_estado = df_estado[
                    df_estado["cedula"] == cedula_seleccionada
                ].iloc[0]

                st.markdown("---")

                col1, col2 = st.columns(2)

                with col1:

                    st.write(
                        f"**Nombre:** {empleado_estado['nombre_completo']}"
                    )

                    st.write(
                        f"**Cédula:** {empleado_estado['cedula']}"
                    )

                    st.write(
                        f"**Cargo:** {empleado_estado['cargo']}"
                    )

                with col2:

                    st.write(
                        f"**Proyecto:** {empleado_estado['proyecto']}"
                    )

                    st.write(
                        f"**Zona:** {empleado_estado['zona']}"
                    )

                    st.write(
                        f"**Estado actual:** {empleado_estado['estado']}"
                    )

                estado_nuevo = st.radio(
                    "Nuevo estado del empleado",
                    [
                        "ACTIVO",
                        "INACTIVO"
                    ],
                    horizontal=True,
                    index=0 if empleado_estado["estado"] == "ACTIVO" else 1,
                    key=f"estado_nuevo_{cedula_seleccionada}"
                )

                if st.button(
                    "💾 Guardar estado",
                    use_container_width=True,
                    key="guardar_estado_empleado"
                ):

                    try:

                        with engine.begin() as conn:

                            conn.execute(
                                text("""
                                    UPDATE empleados
                                    SET estado = :estado
                                    WHERE cedula = :cedula
                                """),
                                {
                                    "estado": estado_nuevo,
                                    "cedula": cedula_seleccionada
                                }
                            )

                        if estado_nuevo == "ACTIVO":

                            st.session_state["mensaje_estado_empleado"] = (
                                "success",
                                "✅ El empleado se ha activado correctamente."
                            )

                        else:

                            st.session_state["mensaje_estado_empleado"] = (
                                "warning",
                                "⚠️ El empleado se ha inactivado correctamente. Ya no podrá registrar asistencias."
                            )
                        
                        st.rerun()

                    except Exception as e:

                        st.error(
                            f"❌ Error actualizando estado: {e}"
                        )
    # =========================================================
    # TAB ASISTENCIAS
    # =========================================================

    with tab_asistencias:

        st.subheader(
            "📋 Reportes Asistencia"
        )

        with engine.begin() as conn:

            df_formaciones = pd.read_sql(
                text("""
                    SELECT
                        id,
                        nombre_formacion,
                        fecha_asistencia
                    FROM formaciones
                    ORDER BY id DESC
                """),
                conn
            )

        opciones = {
            f"{row.id} - {row.nombre_formacion}": row.id
            for row in df_formaciones.itertuples()
        }

        if not opciones:

            st.warning(
                "⚠️ No existen formaciones registradas."
            )

            st.info(
                "Cree una formación para visualizar reportes."
            )

            st.stop()

        seleccion = st.selectbox(
            "Seleccione formación",
            list(opciones.keys())
        )

        id_seleccionado = opciones[seleccion]
        # =====================================================
        # ACTUALIZAR REPORTE
        # =====================================================

        if st.button(
            "🔄 Actualizar reporte",
            use_container_width=True
        ):

            st.rerun()
        with engine.begin() as conn:

            df_asistencias = pd.read_sql(
                text("""
                    SELECT
                        a.cedula,
                        a.nombre_completo,
                        a.cargo,
                        a.proyecto,
                        a.zona,
                        a.formador,
                        f.nombre_formacion,
                        f.fecha_asistencia,
                        a.clasificacion_formacion,
                        a.tipo_formacion,
                        a.autoriza_datos,
                        a.respuesta_1,
                        a.resultado_1,
                        a.respuesta_2,
                        a.resultado_2,
                        a.respuesta_3,
                        a.resultado_3,
                        a.respuesta_4,
                        a.resultado_4,
                        a.respuesta_5,
                        a.resultado_5,
                        a.puntaje,
                        a.fecha_registro
                    FROM asistencias a
                    INNER JOIN formaciones f
                        ON a.id_formacion = f.id
                    WHERE a.id_formacion = :id_formacion
                    ORDER BY a.fecha_registro DESC
                """),
                conn,
                params={
                    "id_formacion": id_seleccionado
                }
            )

        st.dataframe(
            df_asistencias,
            use_container_width=True
        )

        # =====================================================
        # KPIS
        # =====================================================

        col1, col2 = st.columns([1, 3])

        with col1:

            st.metric(
                "👥 Total Asistencias",
                len(df_asistencias)
            )

        with col2:

            nombre_kpi = seleccion.split("-", 1)[1].strip()

            st.html(f"""
            <div style="
                background:#f0fdf4;
                border-radius:14px;
                padding:14px 18px;
                border:1px solid #e2e8f0;
            ">
                <div style="
                    font-size:13px;
                    color:#64748b;
                    font-weight:600;
                    margin-bottom:6px;
                ">
                    📚 Formación
                </div>

                <div style="
                    font-size:18px;
                    color:#0f172a;
                    font-weight:800;
                    line-height:1.25;
                ">
                    {nombre_kpi}
                </div>
            </div>
            """)

        # =====================================================
        # EXPORTAR EXCEL
        # =====================================================

        if not df_asistencias.empty:

            output = BytesIO()

            with pd.ExcelWriter(
                output,
                engine="openpyxl"
            ) as writer:

                df_asistencias.to_excel(
                    writer,
                    index=False,
                    sheet_name="Reporte"
                )

                worksheet = writer.sheets["Reporte"]

                for cell in worksheet[1]:

                    cell.font = Font(
                        bold=True,
                        color="FFFFFF"
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                    cell.fill = PatternFill(
                        start_color="166534",
                        end_color="166534",
                        fill_type="solid"
                    )

                for column_cells in worksheet.columns:

                    length = max(
                        len(str(cell.value))
                        if cell.value else 0
                        for cell in column_cells
                    )

                    worksheet.column_dimensions[
                        get_column_letter(
                            column_cells[0].column
                        )
                    ].width = length + 5

                total_filas = len(df_asistencias) + 1
                total_columnas = len(df_asistencias.columns)

                rango = (
                    f"A1:"
                    f"{get_column_letter(total_columnas)}"
                    f"{total_filas}"
                )

                tabla = Table(
                    displayName="TablaAsistencias",
                    ref=rango
                )

                estilo = TableStyleInfo(
                    name="TableStyleMedium9",
                    showRowStripes=True,
                    showColumnStripes=False
                )

                tabla.tableStyleInfo = estilo

                worksheet.add_table(tabla)

            output.seek(0)

            st.download_button(
                label="📥 Descargar Reporte Excel",
                data=output,
                file_name=(
                    f"Reporte_Asistencia_"
                    f"{id_seleccionado}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
                use_container_width=True
            )